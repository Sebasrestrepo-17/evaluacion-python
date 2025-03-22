import openpyxl

libro = openpyxl.Workbook()

libro.save("Mi-primer-excel")

print("Archivo creado")

# Crear un archivo nuevo
libro = openpyxl.Workbook()

# Elegir la hoja donde escribir (la que ya viene)
hoja = libro.active

# Escribir en celdas como si fuera una cuadrícula
hoja["A1"] = "Hola"
hoja["B1"] = "Mundo"

# Guardar el archivo
libro.save("mi_primer_excel.xlsx")

print("¡Texto agregado!")

# Abrir el archivo que ya hicimos
libro = openpyxl.load_workbook("mi_primer_excel.xlsx")

# Elegir la hoja
hoja = libro.active

# Leer lo que está en A1 y B1
print(hoja["A1"].value)  # Muestra "Hola"
print(hoja["B1"].value)  # Muestra "Mundo"

# No necesitamos guardarlo porque solo leímos


# Abrir el archivo
libro = openpyxl.load_workbook("mi_primer_excel.xlsx")
hoja = libro.active

# Cambiar lo que está en B1
hoja["B1"] = "Amigos"

# Guardar el cambio
libro.save("mi_primer_excel.xlsx")

print("¡Cambio hecho!")

